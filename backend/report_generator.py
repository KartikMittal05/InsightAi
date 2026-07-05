import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class BusinessInsightsReportGenerator:
    """Generate downloadable business insights reports in Excel and PDF formats"""
    
    @staticmethod
    def generate_excel_report(analysis_data, dataset_type_info=None):
        """Generate comprehensive Excel report with business insights"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        title_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        title_font = Font(bold=True, size=11)
        
        value_font = Font(size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "Business Insights Report"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        if dataset_type_info:
            ws['A3'] = f"Dataset: {dataset_type_info.get('name', 'Unknown')}"
            ws['A3'].font = Font(size=10)
        
        # KPIs Section
        ws['A5'] = "Key Performance Indicators (KPIs)"
        ws['A5'].font = title_font
        ws['A5'].fill = title_fill
        ws.merge_cells('A5:D5')
        
        row = 6
        kpis = analysis_data.get('kpis', {})
        
        kpi_metrics = [
            ('Total Revenue', 'total_revenue', '${:,.2f}'),
            ('Total Customers', 'total_customers', '{:,}'),
            ('Average Order Value', 'avg_order_value', '${:,.2f}'),
            ('Growth Rate', 'growth_rate', '{:.2f}%')
        ]
        
        for metric_name, metric_key, format_str in kpi_metrics:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            
            value = kpis.get(metric_key, 0)
            formatted_value = format_str.format(value)
            ws[f'B{row}'] = formatted_value
            ws[f'B{row}'].font = value_font
            
            for col in ['A', 'B']:
                ws[f'{col}{row}'].border = border
            
            row += 1
        
        # Customer Segments
        ws['A12'] = "Customer Segments"
        ws['A12'].font = title_font
        ws['A12'].fill = title_fill
        ws.merge_cells('A12:D12')
        
        segment_counts = analysis_data.get('segment_counts', {})
        revenue_by_segment = analysis_data.get('revenue_by_segment', {})
        
        row = 13
        ws[f'A{row}'] = "Segment"
        ws[f'B{row}'] = "Customer Count"
        ws[f'C{row}'] = "Revenue"
        ws[f'D{row}'] = "% of Total Revenue"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
        
        row += 1
        total_revenue = kpis.get('total_revenue', 0)
        
        for segment, count in segment_counts.items():
            revenue = revenue_by_segment.get(segment, 0)
            pct = (revenue / total_revenue * 100) if total_revenue > 0 else 0
            
            ws[f'A{row}'] = segment
            ws[f'B{row}'] = count
            ws[f'C{row}'] = revenue
            ws[f'D{row}'] = f'{pct:.1f}%'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border
            
            ws[f'C{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Trends Section
        ws['A20'] = "Monthly Trends"
        ws['A20'].font = title_font
        ws['A20'].fill = title_fill
        ws.merge_cells('A20:D20')
        
        trends = analysis_data.get('trends', {})
        trend_labels = trends.get('labels', [])
        revenue_trend = trends.get('revenue_trend', [])
        customer_trend = trends.get('customer_trend', [])
        
        if trend_labels:
            row = 21
            ws[f'A{row}'] = "Month"
            ws[f'B{row}'] = "Revenue"
            ws[f'C{row}'] = "Unique Customers"
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            
            row += 1
            for i, month in enumerate(trend_labels):
                ws[f'A{row}'] = month
                ws[f'B{row}'] = revenue_trend[i] if i < len(revenue_trend) else 0
                ws[f'C{row}'] = customer_trend[i] if i < len(customer_trend) else 0
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                
                ws[f'B{row}'].number_format = '$#,##0.00'
                row += 1
        
        # Lifecycle Analysis
        ws['A35'] = "Customer Lifecycle Status"
        ws['A35'].font = title_font
        ws['A35'].fill = title_fill
        ws.merge_cells('A35:C35')
        
        lifecycle_labels = analysis_data.get('lifecycle_labels', [])
        lifecycle_values = analysis_data.get('lifecycle_values', [])
        
        if lifecycle_labels:
            row = 36
            ws[f'A{row}'] = "Status"
            ws[f'B{row}'] = "Customer Count"
            ws[f'C{row}'] = "% of Total"
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            
            row += 1
            total_customers = kpis.get('total_customers', 1)
            
            for label, value in zip(lifecycle_labels, lifecycle_values):
                pct = (value / total_customers * 100) if total_customers > 0 else 0
                
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'C{row}'] = f'{pct:.1f}%'
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                
                row += 1
        
        # Behavior Analysis
        ws['A50'] = "Customer Behavior"
        ws['A50'].font = title_font
        ws['A50'].fill = title_fill
        ws.merge_cells('A50:C50')
        
        behavior_labels = analysis_data.get('behavior_labels', [])
        behavior_values = analysis_data.get('behavior_values', [])
        
        if behavior_labels:
            row = 51
            ws[f'A{row}'] = "Behavior Type"
            ws[f'B{row}'] = "Count"
            ws[f'C{row}'] = "% of Total"
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            
            row += 1
            total_count = sum(behavior_values) if behavior_values else 1
            
            for label, value in zip(behavior_labels, behavior_values):
                pct = (value / total_count * 100) if total_count > 0 else 0
                
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'C{row}'] = f'{pct:.1f}%'
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Create RFM Customers sheet
        rfm_customers = analysis_data.get('rfm_customers', [])
        if rfm_customers:
            ws_rfm = wb.create_sheet("Top Customers")
            
            # Add title
            ws_rfm['A1'] = "Top 200 Customers by Value"
            ws_rfm['A1'].font = Font(bold=True, size=12, color="1F4E78")
            
            # Convert to DataFrame and write
            df_rfm = pd.DataFrame(rfm_customers)
            for r_idx, row_data in enumerate(dataframe_to_rows(df_rfm, index=False, header=True), 3):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws_rfm.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:
                        cell.font = header_font
                        cell.fill = header_fill
                    cell.border = border
            
            ws_rfm.column_dimensions['A'].width = 15
            ws_rfm.column_dimensions['B'].width = 15
            ws_rfm.column_dimensions['C'].width = 15
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def generate_csv_summary(analysis_data):
        """Generate summary CSV with key metrics and trends"""
        
        summary_data = {
            'Metric': [],
            'Value': []
        }
        
        kpis = analysis_data.get('kpis', {})
        for metric, value in kpis.items():
            summary_data['Metric'].append(metric.replace('_', ' ').title())
            summary_data['Value'].append(value)
        
        df = pd.DataFrame(summary_data)
        
        output = io.StringIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        return output.getvalue()
